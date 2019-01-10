import json , urllib.request, os, sys
import datetime
import time
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range

## --------------------
## Function definitions
## --------------------
      
def print_menu():       ## Your menu design here
    print (30 * "-" , "MENU" , 30 * "-")
    print ("1. Validate VAT related Partner /Alex/")
    print ("2. Validate nonVAT related Partner /Alex/")
    print ("3. Validate Deferred VAT (Vendors only) /Alex/")
    print ("4. Validate VAT Split (Vendors / Customers if needed) /Alex/")
    print ("5. Search for VAT split subjects /Misi/")
    print ("6. Exit")
    print (67 * "-")

def print_datemenu():
    print (60 * "-")
    print ("Query status of vendors as of today?")
    today = input("Press 'y' or 'n' and Enter: ")

    if today == "y":
        ##print (today)
        today_date = datetime.date.today()
        ##print (today_date)
        querydate = today_date.strftime("%Y-%m-%d")
        ##print (querydate)

    else:
        print_dateselector()
        querydate = input('Query date: ')

    return querydate


def print_dateselector():
    print ("Enter query date as 'YYYY-MM-DD'")



## ------------------------
## Determining query option
## ------------------------

print_menu()        ## Displays menu

choice = input("Enter your choice [1-5]: ")

choice = int(choice)        ## Convert string to int type

## --------------------
## Source file path
## --------------------

        ##file_location = "C:\\Users\\mikesm01\\Desktop\\VAT_validation.xlsx"
        ##file_location = "C:\\Users\\" + os.getenv('username') +"\\Desktop\\VAT_validation.xlsx"
## print (sys.path[0])
file_location = os.getcwd() + "\\VAT_validation.xlsx"
## print (file_location)
print (os.getcwd())

wb = load_workbook(file_location)
ws = wb.worksheets[0]
i=2  
j=501


## --------------------
## Running the query
## --------------------

querydate = print_datemenu()

print("Retrieving information for ",ws.max_row-1," partners, as per ",querydate)

if choice==1:

    print ("Validate VAT related parners")      ## action aplicable for nonVAT check
    
elif choice==2:

    print ("Validate nonVAT related parners")       ## action aplicable for VAT check

elif choice==3:

    print ("Validate Deferred VAT (Vendors only)")      ## action aplicable for Deferred VAT check

elif choice==4:

    print ("Validate nonVAT related parners")       ## action aplicable for VAT check

    

elif choice==5:

    print("Validate VAT split status for vendors")
    
    print("VAT ID$NAME$ADDRESS$VAT SPLIT START DATE$VAT SPLIT END DATE$VAT CODE") #Header EN

elif choice==6:

    print ("Exit Menu has been selected")
    ## You can add your code or functions here
    loop=False # This will make the while loop to end as not value of loop is set to False
                                
else:

    # Any integer inputs other than values 1-5 we print an error message
    print("Wrong option selection. Enter any key to try again..")

loop=True

while loop:

    while j<=ws.max_row+1:

            newConditions = []

            ## printing out variables for control

            ## print ("Suppliers from #",i-1," to #",j-1)
            
            for row in range(i, j):
                            
                            newConditions.append({"cui": ws['C' + str(row)].value,"data": querydate})
                            i=i+1
            
            params = json.dumps(newConditions).encode('utf8')

            conditionsSetURL = 'https://webservicesp.anaf.ro:/PlatitorTvaRest/api/v3/ws/tva'

            req = urllib.request.Request(conditionsSetURL, data=params,headers={'content-type': 'application/json'})
        
            response = urllib.request.urlopen(req)
        
            r = response.read().decode('utf8')

            #print(r)
        
            #print(type(r))
        
            partners=json.loads(r)

            #print(partners)
                                
            if choice==1:

                                    print("cui $ scpTVA $ data sfarsit ScpTVA $ data anul imp ScpTVA $ denumire $ mesaj ScpTVA") #ANAF Web service fields

                                    #print("VAT ID$IS TAX PAYER$ADDRESS$VAT SPLIT START DATE$VAT SPLIT END DATE$VAT CODE") #Header EN

                                    for item in partners['found']:
              
                                          if item['mesaj_ScpTVA']=='neplatitor de TVA la data cautata':
                      
                                                print (item['cui'],"$",item['scptva'],"$",item['data_sfarsit_ScpTVA'],"$","$",item['data_anul_imp_ScpTVA'],"$",item['denumire'],"$", item['mesaj_ScpTVA'])
                                      
                                    for item in partners ['notfound']:

                                                print (item['cui'],"$",item['scpTVA'],"$",item['data_sfarsit_ScpTVA'],"$","$",item['data_anul_imp_ScpTVA'],"$",item['denumire'],"$", item['mesaj_ScpTVA'])

                                   
            elif choice==2:  

                                    print("cui $ scpTVA $ data sfarsit ScpTVA $ data anul imp ScpTVA $ denumire $ mesaj ScpTVA") #ANAF Web service fields

                                    for item in partners ['found']:
                                                  
                                          #if item['mesaj']=='platitor de TVA la data cautata':
                                                          
                                                   print (item['cui'],"$",item['scpTVA'],"$",item['data_sfarsit_ScpTVA'],"$",item['data_anul_imp_ScpTVA'],"$",item['denumire'],"$", item['mesaj_ScpTVA'])

                                   
            elif choice==3:

                                    print("cui $ scpTVA $ data sfarsit ScpTVA $ data inceput ScpTVA $ denumire $ mesaj ScpTVA") #ANAF Web service fields

                                    for item in partners ['found']:        
                                                  
                                                   print (item['cui'],"$",item['scpTVA'],"$",item['data_sfarsit_ScpTVA'],"$",item['data_inceput_ScpTVA'],"$",item['denumire'],"$", item['mesaj_ScpTVA'])
                                                   

            elif choice==4:

                                    print("cui $ adresa  $ tipActTvaInc $ scpTVA $ data sfarsit ScpTVA $ data inceput ScpTVA $ denumire $ mesaj ScpTVA $ status SplitTVA") #ANAF Web service fields

                                    for item in partners ['found']:
                                                          
                                                   print (item['cui'],"$",item['adresa'],"$",item['tipActTvaInc'],"$","item['scpTVA']","$",item['data_sfarsit_ScpTVA'],"$",item['data_inceput_ScpTVA'],"$",item['denumire'],"$", item['mesaj_ScpTVA'],"$",item['statusSplitTVA'])


            elif choice==5:

                                    print("cui $ denumire $ adresa  $ dataInceputSplitTVA $ dataAnulareSplitTVA $ cui") #ANAF Web service fields

                                    for item in partners ['found']:
                                                                  
                                            print (item['cui'],"$",item['denumire'],"$",item['adresa'],"$",item['dataInceputSplitTVA'],"$",item['dataAnulareSplitTVA'],"$",str(item['cui']))

            ## printing out variables for control

            ## print(i,j,ws.max_row+1,loop)

            time.sleep(2)

            if j+500>ws.max_row+1:

                    if j==ws.max_row+1:

                            j=ws.max_row+2

                            loop=False

                    else: j=ws.max_row+1
                                            
            else:
                                                
                    j+=500

                              
print(155*'-')
print('The program is finished. Please copy the results into Excel and save it for further investigation. The program will close down automatically after 5 minutes.')                         
print(155*'-')

time.sleep(300)


#ANAF REFERENCE

#"cui": codul fiscal, VAT ID
#"data": data_pt_care_se_efectueaza_cautarea, Date of Query
#"denumire": denumire, Name
#"adresa": adresa, Address
#"scpTVA": true (pentru platitor in scopuri de tva) / false (in cazul in care nu e platitor  in scopuri de TVA la data cautata), Is Tax Payer? (True / False)
#"data_inceput_ScpTVA": 
#"data_sfarsit_ScpTVA": " ",
#"data_anul_imp_ScpTVA": " ",
#"mesaj_ScpTVA": "---MESAJ:(ne)platitor de TVA la data cautata---",
#"dataInceputTvaInc": " ",
#"dataSfarsitTvaInc": " ",
#"dataActualizareTvaInc": " ",
#"dataPublicareTvaInc": " ",
#"tipActTvaInc": " ",
#"statusTvaIncasare":  true -pentru platitor TVA la incasare/ false in cazul in care nu e platitor de TVA la incasare la data cautata
#"dataInactivare": " ",
#"dataReactivare": " ",
#"dataPublicare": " ",
#"dataRadiere": " ",
#"statusInactivi": true -pentru inactiv / false in cazul in care nu este inactiv la data cautata
#"dataInceputSplitTVA": " ",
#"dataAnulareSplitTVA": " ",
#"statusSplitTVA": true -aplica plata defalcata a Tva / false - nu aplica plata defalcata a Tva la data cautata
